"""
Sales Navigator source plugin.

Reads the cleaned CG Sales Nav Leads Excel file, groups leads by
account owner (seller), and cross-references with Demandbase signals
to flag leads at accounts showing active intent this week.
"""

from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

SIGNAL_TYPE_META = {
    "top_leads": {
        "label": "Sales Nav Top Leads",
        "short_label": "Top Leads",
        "color": "teal",
        "source": "salesnav",
        "description": (
            "Your highest-priority contacts from LinkedIn Sales Navigator, "
            "ranked by account fit quality. Leads at accounts with active "
            "Demandbase intent signals this week are highlighted — these are "
            "the people to reach out to right now."
        ),
        "display_columns": [
            {"key": "name", "label": "Name"},
            {"key": "title", "label": "Title"},
            {"key": "company", "label": "Company"},
            {"key": "city", "label": "City"},
            {"key": "fit_score", "label": "Fit"},
            {"key": "priority", "label": "Priority"},
        ],
    },
}

FIT_RANK = {"Excellent Fit": 0, "Good Fit": 1, "Potential Fit": 2, "Poor Fit": 3}
PRIORITY_RANK = {"High": 0, "Medium": 1, "Low": 2}

INTENT_SIGNAL_TYPES = ("mqa_new", "hvp", "hvp_all", "all_mqa")


def _sort_key(lead: dict) -> tuple:
    """Sort leads: intent-active first, then by fit, then priority."""
    return (
        0 if lead["intent_active"] else 1,
        FIT_RANK.get(lead["fit_score"], 9),
        PRIORITY_RANK.get(lead["priority"], 9),
    )


def load(filepath: Path, signals_by_seller: dict) -> dict[str, list[dict]]:
    """
    Load Sales Nav leads from the cleaned Excel file.

    Cross-references each lead's Matched Account against the seller's
    Demandbase signals to set intent_active and intent_types.

    Returns {seller_name: [lead_dicts]} for sellers that have leads.
    """
    if openpyxl is None:
        print("  WARNING: openpyxl not installed — skipping Sales Nav leads")
        return {}

    if not filepath.exists():
        print(f"  WARNING: Sales Nav file not found: {filepath} — skipping")
        return {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Leads (Cleaned)"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    leads_by_seller: dict[str, list[dict]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        owner = (row[col_idx["Account Owner"]] or "").strip()
        if not owner:
            continue

        first = (row[col_idx["First Name"]] or "").strip()
        last = (row[col_idx["Last Name"]] or "").strip()
        name = f"{first} {last}".strip()

        lead = {
            "name": name,
            "title": (row[col_idx["Title"]] or "").strip(),
            "company": (row[col_idx["Company"]] or "").strip(),
            "city": (row[col_idx["City"]] or "").strip(),
            "fit_score": (row[col_idx["Fit Score"]] or "").strip(),
            "priority": (row[col_idx["Priority"]] or "").strip(),
            "matched_account": (row[col_idx["Matched Account"]] or "").strip(),
            "industry": (row[col_idx["Industry"]] or "").strip(),
            "intent_active": False,
            "intent_types": [],
        }

        if owner not in leads_by_seller:
            leads_by_seller[owner] = []
        leads_by_seller[owner].append(lead)

    wb.close()

    # Cross-reference with Demandbase signals for intent flags
    for seller_name, leads in leads_by_seller.items():
        seller_accounts: dict[str, list[str]] = {}
        seller_signals = signals_by_seller.get(seller_name, {})

        for sig_type in INTENT_SIGNAL_TYPES:
            for sig in seller_signals.get(sig_type, []):
                acct = (sig.get("account") or "").strip().lower()
                if acct:
                    if acct not in seller_accounts:
                        seller_accounts[acct] = []
                    if sig_type not in seller_accounts[acct]:
                        seller_accounts[acct].append(sig_type)

        for lead in leads:
            matched = lead["matched_account"].lower()
            if matched and matched in seller_accounts:
                lead["intent_active"] = True
                lead["intent_types"] = seller_accounts[matched]

        leads.sort(key=_sort_key)

    return leads_by_seller
